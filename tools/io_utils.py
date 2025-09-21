#!/usr/bin/env python3
"""
I/O utilities to load tabular data from CSV or XLSX without losing columns.

Goals
- Accept .csv or .xlsx seamlessly
- Preserve all columns and rows (no drops)
- Represent values as strings by default to avoid implicit type coercion

Notes
- For .xlsx, this uses openpyxl if available. If not installed, raises
  an explicit error suggesting installation or exporting to CSV.
"""
from __future__ import annotations

import csv
import os
from typing import Iterable, List, Dict, Any


def _read_csv(path: str) -> List[Dict[str, str]]:
    # Try to sniff delimiter; default to comma, then semicolon
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            # Heuristic: prefer semicolon if there are many
            delim = ";" if sample.count(";") > sample.count(",") else ","
            dialect = csv.excel
            dialect.delimiter = delim  # type: ignore[attr-defined]
        reader = csv.DictReader(fh, dialect=dialect)
        rows: List[Dict[str, str]] = []
        for r in reader:
            # Ensure all values as strings (preserve empty as "")
            rows.append({(k or "").strip(): ("" if v is None else str(v)) for k, v in r.items()})
        return rows


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    # Avoid scientific notation for floats
    if isinstance(value, float):
        s = ("%f" % value).rstrip("0").rstrip(".")
        return s if s else "0"
    return str(value)


def _read_xlsx(path: str, sheet: str | int | None = None) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Soporte XLSX requiere 'openpyxl'. Instala con: pip install openpyxl\n"
            "O exporta el archivo a CSV.\n"
            f"Archivo: {path}"
        ) from e

    wb = load_workbook(path, data_only=True, read_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    elif isinstance(sheet, str):
        ws = wb[sheet]
    else:
        ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    headers = [(_stringify(h)).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(headers_raw)]

    rows: List[Dict[str, str]] = []
    for row in rows_iter:
        values = [_stringify(v) for v in row]
        # Pad if needed
        if len(values) < len(headers):
            values += [""] * (len(headers) - len(values))
        item = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(item)
    return rows


def load_rows(path: str, *, sheet: str | int | None = None) -> List[Dict[str, str]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_csv(path)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _read_xlsx(path, sheet=sheet)
    raise ValueError(f"Formato no soportado: {ext}. Usa CSV o XLSX.")

